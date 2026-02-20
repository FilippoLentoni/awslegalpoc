"""Seed an evaluation dataset into Langfuse from an XLSX test set.

Loads questions from each sheet of the XLSX file and creates Langfuse
dataset items with input (Domanda) and expected output (Risposta).

Re-running this script upserts items (same ID = overwrite, not duplicate).

Usage:
    # Download test set from S3
    aws s3 cp s3://materialpoc/knowledge-base/test_set.xlsx /tmp/test_set.xlsx

    # Seed dataset
    set -a && source .env && set +a
    python3.11 scripts/seed_langfuse_dataset.py --xlsx /tmp/test_set.xlsx
"""

import argparse
import hashlib
import sys
import os
import time

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from openpyxl import load_workbook

from core.langfuse_client import get_langfuse_client

DATASET_NAME = "italian-legal-eval"

MAX_RETRIES = 5


def _call_with_retry(fn, *args, **kwargs):
    """Call a function with exponential backoff on 429 errors."""
    for attempt in range(MAX_RETRIES):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            if "429" in str(e) and attempt < MAX_RETRIES - 1:
                wait = 2 ** (attempt + 1)
                print(f"    Rate limited, waiting {wait}s...")
                time.sleep(wait)
            else:
                raise


def load_items_from_xlsx(xlsx_path: str) -> list:
    """Load test items from XLSX file with multiple sheets."""
    wb = load_workbook(xlsx_path, data_only=True)
    items = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Build header index from first row
        headers = {}
        for idx, cell in enumerate(ws[1]):
            if cell.value:
                headers[cell.value.strip()] = idx

        domanda_col = headers.get("Domanda")
        risposta_col = headers.get("Risposta (quella che vorremmo che il bot fornisse)")
        tipologia_col = headers.get("Tipologia")
        riferimenti_col = headers.get("Riferimenti")
        n_col = headers.get("N.")

        if domanda_col is None or risposta_col is None:
            print(f"  Skipping sheet '{sheet_name}': missing required columns")
            continue

        sheet_count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = list(row)
            domanda = cells[domanda_col].value
            risposta = cells[risposta_col].value

            # Skip empty rows
            if not domanda or not str(domanda).strip():
                continue
            if not risposta or not str(risposta).strip():
                continue

            tipologia = cells[tipologia_col].value if tipologia_col is not None else None
            riferimenti = cells[riferimenti_col].value if riferimenti_col is not None else None
            n = cells[n_col].value if n_col is not None else None

            items.append({
                "input": {"input": str(domanda).strip()},
                "expected_output": str(risposta).strip(),
                "metadata": {
                    "domain": sheet_name,
                    "tipologia": str(tipologia) if tipologia else None,
                    "riferimenti": str(riferimenti) if riferimenti else None,
                    "question_number": int(n) if n else None,
                },
            })
            sheet_count += 1

        print(f"  Sheet '{sheet_name}': {sheet_count} items loaded")

    return items


def main():
    parser = argparse.ArgumentParser(description="Seed Langfuse eval dataset from XLSX")
    parser.add_argument("--xlsx", required=True, help="Path to test_set.xlsx file")
    parser.add_argument("--dataset", default=DATASET_NAME, help="Langfuse dataset name")
    parser.add_argument("--max-items", type=int, default=None,
                        help="Limit to first N items (in sheet order)")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: File not found: {args.xlsx}")
        sys.exit(1)

    lf = get_langfuse_client()
    if not lf:
        print("ERROR: Langfuse client not configured.")
        sys.exit(1)

    # Load items from XLSX
    print(f"Loading items from: {args.xlsx}")
    items = load_items_from_xlsx(args.xlsx)
    print(f"Total items loaded: {len(items)}")

    if args.max_items and args.max_items < len(items):
        items = items[:args.max_items]
        print(f"Truncated to first {args.max_items} items")

    if not items:
        print("ERROR: No items loaded from XLSX.")
        sys.exit(1)

    # Create dataset
    try:
        lf.create_dataset(
            name=args.dataset,
            description="Italian notarial law evaluation dataset",
        )
        print(f"Created dataset '{args.dataset}'")
    except Exception:
        print(f"Dataset '{args.dataset}' already exists, updating items...")

    # Archive all existing items (replacing dataset contents)
    try:
        existing = lf.get_dataset(args.dataset)
        for old_item in existing.items:
            if getattr(old_item, "status", "ACTIVE") != "ARCHIVED":
                _call_with_retry(
                    lf.create_dataset_item,
                    dataset_name=args.dataset,
                    id=old_item.id,
                    input=old_item.input,
                    status="ARCHIVED",
                )
                print(f"  Archived old item: {old_item.id}")
    except Exception:
        pass

    # Add new items (auto-generated IDs)
    for i, item in enumerate(items):
        _call_with_retry(
            lf.create_dataset_item,
            dataset_name=args.dataset,
            input=item["input"],
            expected_output=item["expected_output"],
            metadata=item["metadata"],
        )
        print(f"  [{i + 1}/{len(items)}] {item['input']['input'][:60]}")
        # Rate limit: delay every 5 items to avoid 429 errors
        if (i + 1) % 5 == 0:
            time.sleep(3)

    lf.flush()
    print(f"\nDone! {len(items)} items added to dataset '{args.dataset}'.")


if __name__ == "__main__":
    main()
